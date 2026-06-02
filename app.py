import streamlit as st
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NS = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}

def tv(node, tag):
    el = node.find(tag, NS) if node is not None else None
    return el.text.strip() if el is not None and el.text else ''

def num(node, tag):
    try: return float(tv(node, tag))
    except: return 0.0

def parse_nfe(xml_bytes):
    try:
        root = ET.fromstring(xml_bytes)
    except:
        return []
    nfe = root.find('.//nfe:infNFe', NS)
    if nfe is None:
        return []
    ide   = nfe.find('nfe:ide',  NS)
    emit  = nfe.find('nfe:emit', NS)
    dest  = nfe.find('nfe:dest', NS)
    ender = emit.find('nfe:enderEmit', NS) if emit is not None else None
    tot   = nfe.find('.//nfe:ICMSTot', NS)
    prot  = root.find('.//nfe:infProt', NS)

    raw = tv(ide, 'nfe:dhEmi')
    try:    data = datetime.fromisoformat(raw).strftime('%d/%m/%Y')
    except: data = raw[:10]

    base = dict(
        chave      = tv(prot, 'nfe:chNFe'),
        nNF        = tv(ide,  'nfe:nNF'),
        serie      = tv(ide,  'nfe:serie'),
        data       = data,
        natureza   = tv(ide,  'nfe:natOp'),
        status     = tv(prot, 'nfe:xMotivo'),
        emit_cnpj  = tv(emit, 'nfe:CNPJ'),
        emit_nome  = tv(emit, 'nfe:xNome'),
        emit_uf    = tv(ender,'nfe:UF'),
        emit_mun   = tv(ender,'nfe:xMun'),
        dest_cnpj  = tv(dest, 'nfe:CNPJ') or tv(dest, 'nfe:CPF'),
        dest_nome  = tv(dest, 'nfe:xNome'),
        vProd      = num(tot, 'nfe:vProd'),
        vDesc      = num(tot, 'nfe:vDesc'),
        vFrete     = num(tot, 'nfe:vFrete'),
        vICMS      = num(tot, 'nfe:vICMS'),
        vPIS       = num(tot, 'nfe:vPIS'),
        vCOFINS    = num(tot, 'nfe:vCOFINS'),
        vNF        = num(tot, 'nfe:vNF'),
        infCpl     = tv(nfe.find('nfe:infAdic', NS), 'nfe:infCpl'),
    )

    rows = []
    for det in nfe.findall('nfe:det', NS):
        prod = det.find('nfe:prod', NS)
        imp  = det.find('nfe:imposto', NS)
        rows.append({**base,
            'item':     det.get('nItem', ''),
            'cod_prod': tv(prod, 'nfe:cProd'),
            'descricao':tv(prod, 'nfe:xProd'),
            'ncm':      tv(prod, 'nfe:NCM'),
            'cfop':     tv(prod, 'nfe:CFOP'),
            'unidade':  tv(prod, 'nfe:uCom'),
            'qtd':      num(prod,'nfe:qCom'),
            'vl_unit':  num(prod,'nfe:vUnCom'),
            'vl_item':  num(prod,'nfe:vProd'),
            'vTotTrib': num(imp, 'nfe:vTotTrib'),
        })
    return rows

def build_excel(all_rows):
    wb = openpyxl.Workbook()
    HDR  = PatternFill('solid', start_color='1F3864')
    HFNT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    ALT  = PatternFill('solid', start_color='DCE6F1')
    BRD  = Border(bottom=Side(style='thin', color='BFBFBF'),
                  right =Side(style='thin', color='BFBFBF'))
    FMT  = '#,##0.00'

    def hdr(ws, cols):
        ws.append(cols)
        for c in range(1, len(cols)+1):
            cl = ws.cell(1, c)
            cl.fill = HDR; cl.font = HFNT
            cl.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cl.border = BRD
        ws.row_dimensions[1].height = 28

    def row_style(ws, r, num_from):
        fill = ALT if r % 2 == 0 else None
        for c in range(1, ws.max_column+1):
            cl = ws.cell(r, c)
            cl.font = Font(name='Arial', size=10); cl.border = BRD
            cl.alignment = Alignment(vertical='center')
            if fill: cl.fill = fill
            if c >= num_from and isinstance(cl.value, float):
                cl.number_format = FMT
                cl.alignment = Alignment(horizontal='right', vertical='center')

    # ── Itens ──
    ws = wb.active; ws.title = 'Itens'
    H = ['Chave NF-e','NF','Série','Data','Natureza','Status',
         'CNPJ Emit.','Emitente','UF','Município',
         'CNPJ/CPF Dest.','Destinatário',
         'Item','Cód.','Descrição','NCM','CFOP','Unid.',
         'Qtd','Vl Unit','Vl Item','Trib. Est.',
         'Vl Total NF','Vl Desc','Vl Frete','ICMS','PIS','COFINS','Inf. Complementares']
    K = ['chave','nNF','serie','data','natureza','status',
         'emit_cnpj','emit_nome','emit_uf','emit_mun',
         'dest_cnpj','dest_nome',
         'item','cod_prod','descricao','ncm','cfop','unidade',
         'qtd','vl_unit','vl_item','vTotTrib',
         'vNF','vDesc','vFrete','vICMS','vPIS','vCOFINS','infCpl']
    hdr(ws, H)
    for r, row in enumerate(all_rows, 2):
        for c, k in enumerate(K, 1):
            ws.cell(r, c, row.get(k, ''))
        row_style(ws, r, 19)
    W = [44,7,6,12,18,20,18,30,4,18,18,30,5,10,34,10,6,6,10,12,12,12,14,10,10,10,9,10,50]
    for i,w in enumerate(W,1): ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'; ws.auto_filter.ref = ws.dimensions

    # ── Resumo por NF ──
    ws2 = wb.create_sheet('Resumo por NF')
    seen = {}
    for row in all_rows:
        if row['chave'] not in seen: seen[row['chave']] = row
    H2 = ['Chave NF-e','NF','Série','Data','Emitente','CNPJ Emit.','Destinatário',
          'Natureza','Status','Vl Produtos','Vl Desc','Vl Frete','ICMS','PIS','COFINS','Vl Total NF','Inf. Complementares']
    hdr(ws2, H2)
    for r, row in enumerate(seen.values(), 2):
        ws2.append([row['chave'],row['nNF'],row['serie'],row['data'],
                    row['emit_nome'],row['emit_cnpj'],row['dest_nome'],
                    row['natureza'],row['status'],
                    row['vProd'],row['vDesc'],row['vFrete'],
                    row['vICMS'],row['vPIS'],row['vCOFINS'],row['vNF'],row['infCpl']])
        row_style(ws2, r, 10)
    W2=[44,6,6,12,30,18,30,18,20,13,10,10,10,9,10,13,50]
    for i,w in enumerate(W2,1): ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = 'A2'; ws2.auto_filter.ref = ws2.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── Streamlit UI ──
st.set_page_config(page_title='Conversor NF-e → Excel', page_icon='📄', layout='centered')
st.title('📄 Conversor NF-e → Excel')
st.caption('Faça upload de um .zip com XMLs de NF-e e baixe a planilha consolidada.')

uploaded = st.file_uploader('Selecione o arquivo ZIP', type='zip')

if uploaded:
    all_rows = []
    erros = []

    with zipfile.ZipFile(uploaded) as z:
        xmls = [n for n in z.namelist() if n.lower().endswith('.xml')]
        st.info(f'{len(xmls)} arquivo(s) XML encontrado(s) no ZIP')

        bar = st.progress(0, text='Processando...')
        for i, name in enumerate(xmls):
            with z.open(name) as f:
                rows = parse_nfe(f.read())
            if rows:
                all_rows.extend(rows)
            else:
                erros.append(name)
            bar.progress((i+1)/len(xmls), text=f'{i+1}/{len(xmls)} — {name.split("/")[-1]}')
        bar.empty()

    if not all_rows:
        st.error('Nenhuma NF-e válida encontrada no ZIP.')
    else:
        nf_count = len({r['chave'] for r in all_rows})
        c1, c2, c3 = st.columns(3)
        c1.metric('NF-e', nf_count)
        c2.metric('Itens', len(all_rows))
        c3.metric('Ignorados', len(erros))

        if erros:
            with st.expander(f'{len(erros)} arquivo(s) ignorado(s)'):
                for e in erros: st.write(e)

        excel_buf = build_excel(all_rows)
        fname = f'nfe_consolidado_{datetime.today().strftime("%Y-%m-%d")}.xlsx'
        st.download_button('⬇️ Baixar Excel', data=excel_buf, file_name=fname,
                           mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                           use_container_width=True)